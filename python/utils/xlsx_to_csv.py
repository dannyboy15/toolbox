from parsons import Table

from openpyxl import load_workbook
import petl
import click
import os


def disect_filename(filename):
    dir, basename = os.path.split(filename)
    name, ext = os.path.splitext(basename)

    return dir, name, ext


def get_sheet_names(filename):
    wb = load_workbook(filename=filename, read_only=True)

    return wb.sheetnames


@click.command()
@click.argument('filename')
def main(filename):
    dir, name, ext = disect_filename(filename)
    dir = f"{dir}/" if dir else ""
    sheets = get_sheet_names(filename)

    if len(sheets) == 1:
        raw_data = petl.fromxlsx(filename)
        tbl = Table(raw_data)

        tbl.to_csv(f"{dir}{name}.csv")

    else:
        for sheet in sheets:
            raw_data = petl.fromxlsx(filename, sheet)
            tbl = Table(raw_data)

            tbl.to_csv(f"{dir}{name}[{sheet}].csv")


if __name__ == '__main__':
    main()
